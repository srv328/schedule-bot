from datetime import datetime, timedelta
import random
import pytz
from aiogram.enums import ParseMode
from config import start_day, start_month
from aiogram.types import Message, FSInputFile
from utils.work_with_db import has_schedule, get_schedule_by_day_offset, get_lesson_by_params_with_user,\
    execute_query, get_existing_lessons
import os
from openpyxl.styles import PatternFill
import openpyxl
from aiogram.types import ReplyKeyboardMarkup, InlineKeyboardMarkup, KeyboardButton, InlineKeyboardButton
from utils.bot_entity import bot


def create_keyboard_markup(buttons, is_inline=False, times=False):
    return InlineKeyboardMarkup(inline_keyboard=buttons) if is_inline else ReplyKeyboardMarkup(keyboard=buttons,
                                                                                               resize_keyboard=True,
                                                                                               one_time_keyboard=times)


def create_text_button(text, callback_data=None):
    return InlineKeyboardButton(text=text, callback_data=callback_data) if callback_data else KeyboardButton(text=text)


def generate_default_share_keyboard():
    return [[InlineKeyboardButton(
        text='Поделиться ссылкой на бота',
        switch_inline_query=f'Ссылка')]]


def generate_share_keyboard():
    return [[InlineKeyboardButton(
        text='Поделиться расписанием',
        switch_inline_query=f'Расписание')]]


day_of_week_dict = {
    1: "понедельник",
    2: "вторник",
    3: "среда",
    4: "четверг",
    5: "пятница",
    6: "суббота",
    7: 'воскресенье'
}

day_translation_form = {
    'monday': 'понедельник',
    'tuesday': 'вторник',
    'wednesday': 'среду',
    'thursday': 'четверг',
    'friday': 'пятницу',
    'saturday': 'субботу',
}

day_translation = {
    'monday': 'понедельник',
    'tuesday': 'вторник',
    'wednesday': 'среда',
    'thursday': 'четверг',
    'friday': 'пятница',
    'saturday': 'суббота',
    'sunday': 'воскресенье'
}


def get_available_lesson_times(user_id, day, week_parity):
    all_lesson_times_query = "SELECT start_time, end_time FROM lesson_times"
    all_lesson_times = [(row[0], row[1]) for row in execute_query(all_lesson_times_query)]

    existing_lesson_times = get_existing_lessons(user_id, day, week_parity)

    available_lesson_times = list(set(all_lesson_times) - set(existing_lesson_times))
    available_lesson_times.sort(key=lambda x: datetime.strptime(x[0], "%H:%M"))

    return available_lesson_times


def get_lesson_info(user_id, subject_number, selected_day, week_parity):
    lesson = get_lesson_by_params_with_user(user_id, subject_number, selected_day, week_parity)[0]

    if lesson:
        print(lesson)
        start_time = lesson[-3]
        end_time = lesson[-2]
        subject_name = lesson[3]
        tutor_name = lesson[2]
        is_practice = "Практика💻" if lesson[4] else "Лекция✏️"
        subject_priority = get_subject_priority_text(lesson[5])
        subject_place = lesson[6]

        message_text = (
            f"<b>Информация о паре:</b>\n"
            f"<b>Предмет:</b> {subject_name}\n"
            f"<b>Преподаватель:</b> {tutor_name}\n"
            f"<b>Тип:</b> {is_practice}\n"
            f"<b>Приоритет:</b> {subject_priority}\n"
            f"<b>Аудитория:</b> {subject_place}\n"
            f"<b>Время:</b> {start_time} - {end_time}\n"
        )
    else:
        message_text = "Пара не найдена."

    return message_text


def get_formatted_date():
    current_time_local = datetime.now(pytz.timezone('Asia/Vladivostok'))
    month_names = ["Января", "Февраля", "Марта", "Апреля", "Мая", "Июня",
                   "Июля", "Августа", "Сентября", "Октября", "Ноября", "Декабря"]
    month_name = month_names[current_time_local.month - 1]
    formatted_date = f"{current_time_local.day} " \
                     f"{month_name} {current_time_local.year} " \
                     f"{current_time_local.hour:02d}-{current_time_local.minute:02d}-{current_time_local.second:02d}"
    return formatted_date


def get_local_time():
    current_time_local = datetime.now(pytz.timezone('Asia/Vladivostok'))
    return current_time_local


def format_time_str(days, hours, minutes, seconds):
    time_str = ""

    if days < 0:
        days = 14 + days

    if days > 0:
        time_str += f"{days} д. "
    if hours > 0:
        time_str += f"{hours} ч. "
    if minutes > 0 or (days == 0 and hours == 0):
        time_str += f"{minutes} мин. "
    if seconds > 0 or (days == 0 and hours == 0 and minutes == 0):
        time_str += f"{seconds} сек."

    return time_str


def get_days_of_week():

    current_time_local = get_local_time()
    start_of_week = current_time_local - timedelta(days=current_time_local.weekday())
    week_dates = []

    for i in range(14):
        current_day = start_of_week + timedelta(days=i)
        if current_day.isoweekday() != 7:
            week_dates.append(current_day.strftime("%d.%m.%Y"))

    return week_dates


def get_next_two_weeks_dates():
    current_time_local = get_local_time()
    start_date = current_time_local.date()
    two_weeks_dates = []

    for i in range(15):  # получаем следующие две недели
        current_day = start_date + timedelta(days=i)
        if current_day.isoweekday() != 7:
            two_weeks_dates.append(current_day.strftime("%d.%m.%Y"))

    return two_weeks_dates


def generate_schedule_response(schedule, day_name):
    response_text = f"<b>Расписание на {day_name}🗓</b>\n"

    book_emojis = ["\U0001F4D9", "\U0001F4D8", " \U0001F4D8", "\U0001F4D5",
                   "\U0001F4D2", "\U0001F4D4", "\U0001F4DA", "\U0001F4D3"]

    for lesson in schedule:
        number_symbol = get_number_of_subject_emoji(lesson[-1])
        subject_name = lesson[3]
        tutor_name = lesson[2]
        is_practice = "Практика💻" if lesson[4] else "Лекция✏️"
        subject_priority = get_subject_priority_text(lesson[5])
        subject_place = lesson[6]
        start_time = lesson[-3]
        end_time = lesson[-2]
        response_text += (
            f"\n<b>Пара</b> {number_symbol} \n"
            f"<b>Предмет:</b> {subject_name} {get_random_book_emoji(book_emojis)}\n"
            f"<b>Преподаватель:</b> {tutor_name} 👨🏻‍🏫👩‍🏫\n"
            f"<b>Тип:</b> {is_practice}\n"
            f"<b>Приоритет:</b> {subject_priority} ❕\n"
            f"<b>Аудитория:</b> {subject_place} 🏫\n"
            f"<b>Время:</b> {start_time} - {end_time} ⏰\n"
        )

    return response_text


def get_random_book_emoji(book_emojis):
    if not book_emojis:
        return "\U0001F4D9"
    random_book = random.choice(book_emojis)
    book_emojis.remove(random_book)
    return random_book


def get_week_title(data):
    return f"<b>Выберите день {'текущей' if data == 'this_week' else 'следующей'} недели</b>\n" \
           f"{'<i>(нечётная)</i>' if get_week_parity() == (data == 'this_week') else '<i>(чётная)</i>'}"


def get_week_parity():
    academic_year_start = datetime(datetime.now().year, start_month, start_day)
    current_date = datetime.now()
    days_since_start = (current_date - academic_year_start).days
    current_week_number = days_since_start // 7 + 1
    week_parity = 0 if current_week_number % 2 == 0 else 1
    return week_parity


def get_subject_priority_text(subject_priority):
    priority_map = {
        0: "Зачет",
        1: "Экзамен",
        2: "Зачет с оценкой",
    }
    return priority_map.get(subject_priority, "Неизвестно")


def get_number_of_subject_emoji(subject_number):
    number_of_subject = {
        1: "\U00000031\U000020E3",
        2: "\U00000032\U000020E3",
        3: "\U00000033\U000020E3",
        4: "\U00000034\U000020E3",
        5: "\U00000035\U000020E3",
        6: "\U00000036\U000020E3",
        7: "\U00000037\U000020E3",
        8: "\U00000038\U000020E3"
    }
    return number_of_subject.get(subject_number, "Недопустимый индекс")


async def create_excel_schedule(message: Message):
    user_id = message.from_user.id
    if not has_schedule(user_id):
        await message.answer(text='<b>Ваше расписание не содержит занятий!</b>\n'
                                  '<b>Сгенерировать файл невозможно.</b>',
                             parse_mode=ParseMode.HTML)
        video = FSInputFile("../media/расписание пустое 2.mp4")
        await message.answer_video(video)
        return

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Удаляем активный лист по умолчанию

    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

    for week_parity in range(2):  # Перебор четных и нечетных недель
        parity_name = "Чётная" if week_parity == 0 else "Нечётная"
        sheet = workbook.create_sheet(title=f"{parity_name} неделя")

        week_bool = False
        headers = ["", "ВРЕМЯ", "ДИСЦИПЛИНА", "АУДИТОРИЯ", "ПРЕПОД", "ВИД", "ОЦЕНИВАНИЕ"]
        sheet.append(headers)
        for day_offset in range(1, 7):  # Перебор дней недели

            book_emojis = ["\U0001F4D9", "\U0001F4D8", " \U0001F4D8", "\U0001F4D5",
                           "\U0001F4D2", "\U0001F4D4", "\U0001F4DA", "\U0001F4D3"]

            day_name = days_of_week[day_offset - 1]
            schedule = get_schedule_by_day_offset(user_id, day_offset, week_parity)

            if schedule:
                week_bool = True
                dark_green_fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')

                sheet.append([f"{day_name.capitalize()}:"])

                sheet.cell(row=sheet.max_row, column=1).fill = dark_green_fill

                for lesson_num, lesson in enumerate(schedule, start=1):
                    row_data = [
                        f"",
                        f"{lesson[-3]} – {lesson[-2]}",
                        f"{lesson[3]}",
                        f"{lesson[6]}",
                        f"{lesson[2]}",
                        f"{'Практика' if lesson[4] else 'Лекция️'}",
                        f"{get_subject_priority_text(lesson[5])}"
                    ]
                    sheet.append(row_data)
        if not week_bool:
            sheet.append(['На этой неделе пар нет.'])

    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        for column in current_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except (AttributeError, TypeError):
                    pass

            adjusted_width = (max_length + 2) * 1.2
            current_sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(f'temp/{user_id}_schedule.xlsx')

    schedule_file = FSInputFile(f'temp/{user_id}_schedule.xlsx', 'Расписание.xlsx')

    await bot.send_document(document=schedule_file,
                            caption=f'<b>Ваше расписание в Excel файле.</b>\n'
                                    f'<i>В файле содержится 2 листа✅</i>',
                            chat_id=user_id,
                            parse_mode=ParseMode.HTML)
    os.remove(f'temp/{user_id}_schedule.xlsx')


async def export_schedule_to_txt(message: Message):
    user_id = message.from_user.id
    if not has_schedule(user_id):
        await message.answer(text='<b>Ваше расписание не содержит занятий!</b>\n'
                                  '<b>Сгенерировать файл невозможно.</b>',
                             parse_mode=ParseMode.HTML)
        video = FSInputFile("../media/расписание пустое.mp4")
        await message.answer_video(video)
        return

    with open(f"temp/{user_id}.txt", "w", encoding="utf-8") as file:
        days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]

        for week_parity in range(2):  # Перебор четных и нечетных недель
            parity_name = "Четная" if week_parity == 0 else "\nНечетная"
            file.write(f"{parity_name} неделя:\n")
            week_bool = False
            for day_offset in range(1, 7):  # Перебор дней недели

                book_emojis = ["\U0001F4D9", "\U0001F4D8", " \U0001F4D8", "\U0001F4D5",
                               "\U0001F4D2", "\U0001F4D4", "\U0001F4DA", "\U0001F4D3"]

                day_name = days_of_week[day_offset - 1]
                schedule = get_schedule_by_day_offset(user_id, day_offset, week_parity)

                if schedule:
                    week_bool = True
                    file.write(f"{day_name.capitalize()}:")
                    for lesson_num, lesson in enumerate(schedule, start=1):
                        file.write(f"\nПара {get_number_of_subject_emoji(lesson[-1])} ")
                        file.write(f"Время: {lesson[-3]} - {lesson[-2]} ⏰ ")
                        file.write(f"Предмет: {lesson[3]} {get_random_book_emoji(book_emojis)} ")
                        file.write(f"Преподаватель: {lesson[2]} 👨🏻‍🏫👩‍🏫 ")
                        file.write(f"Тип: {'Практика💻' if lesson[4] else 'Лекция✏️'} ")
                        file.write(f"Приоритет: {get_subject_priority_text(lesson[5])} ❕ ")
                        file.write(f"Аудитория: {lesson[6]} 🏫 ")
                    file.write("\n")
            if not week_bool:
                file.write('На этой неделе пар нет.\n\n')
    schedule_file = FSInputFile(f'temp/{user_id}.txt', 'Расписание.txt')

    await bot.send_document(document=schedule_file,
                            caption=f'<b>Ваше расписание в текстовом файле.</b>',
                            chat_id=user_id,
                            parse_mode=ParseMode.HTML)
    os.remove(f'temp/{user_id}.txt')


def generate_schedule_statistics_message(even_schedule, odd_schedule):
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    message_lines = ["Статистика <b>вашего</b> расписания 📊:\n\n"]

    has_data_even = any(even_schedule.values())
    has_data_odd = any(odd_schedule.values())

    # чётная неделя
    message_lines.append("<b>Четная неделя:</b>\n")
    if has_data_even:
        for day_index, day_name in enumerate(days_of_week, start=1):
            even_count = even_schedule.get(day_index, 0)
            if even_count > 0:
                message_lines.append(f"{day_name.capitalize()}: {even_count} пар\n")
    else:
        message_lines.append("На этой неделе пар нет.\n")

    message_lines.append("\n")

    # нечётная неделя
    message_lines.append("<b>Нечетная неделя:</b>\n")
    if has_data_odd:
        for day_index, day_name in enumerate(days_of_week, start=1):
            odd_count = odd_schedule.get(day_index, 0)
            if odd_count > 0:
                message_lines.append(f"{day_name.capitalize()}: {odd_count} пар\n")
    else:
        message_lines.append("На этой неделе пар нет.\n")

    if not has_data_even and not has_data_odd:
        return "У вас пока нет занятий в расписании."

    total_even = sum(even_schedule.values())
    total_odd = sum(odd_schedule.values())
    total_all = total_even + total_odd

    message_lines.extend([
        "\n",
        f"<b>Итого за 2 недели:</b>\n"
        f"<i>Чётная неделя:</i> {total_even} пар\n"
        f"<i>Нечётная неделя:</i> {total_odd} пар\n"
        f"{'-' * 40}\n",
        f"В ДВФУ вы проведёте {total_all * 1.5} часов ⏳\n"
    ])

    return "".join(message_lines)


def extract_unique_code(text):
    return text.split()[1] if len(text.split()) > 1 else None
